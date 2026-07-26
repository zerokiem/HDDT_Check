#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
checker_web.py  —  Web-adapted InvoiceChecker v8
Chạy headless, nhận callback tiến độ, trả kết quả dạng list[dict].
"""

import os
import re
import time
import zipfile
import warnings
import logging
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from input_utils import read_input_excel

warnings.filterwarnings('ignore')
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'
logging.getLogger('tensorflow').setLevel(logging.FATAL)

# ── ddddocr ──────────────────────────────────────────────────
try:
    import ddddocr as _ddddocr_mod
    _OCR = _ddddocr_mod.DdddOcr(show_ad=False)
except Exception as e:
    _OCR = None
    logging.warning(f'[checker_web] ddddocr unavailable: {e}')

# ── Playwright ───────────────────────────────────────────────
from playwright.sync_api import sync_playwright, TimeoutError as PwTimeout  # noqa

RESULT_COLOR = {
    'YES':         'C6EFCE',
    'NO':          'FFC7CE',
    'CAPTCHA_ERR': 'FFEB9C',
    'UNKNOWN':     'D9D9D9',
    'ERROR':       'F4B942',
}

OUTPUT_COLUMNS = [
    ('STT',            'STT'),
    ('MST',            'MST'),
    ('Ten_Cty',        'Tên công ty'),
    ('So_HD',          'Số hóa đơn'),
    ('Ky_hieu',        'Ký hiệu HĐ'),
    ('Tien_thue',      'Tiền thuế'),
    ('Thanh_toan',     'Tiền thanh toán'),
    ('Loai_HD',        'Loại hóa đơn'),
    ('Ngay',           'Ngày hóa đơn'),
    ('Ngay_check',     'Ngày kiểm tra'),
    ('Ket_qua',        'Kết quả'),
    ('Mo_ta',          'Mô tả kết quả'),
    ('So_lan_captcha', 'Số lần captcha'),
]


# ════════════════════════════════════════════════════════════
class InvoiceCheckerWeb:
    """
    Tham số:
      input_path     — đường dẫn file Excel đầu vào
      output_dir     — thư mục lưu Excel kết quả
      screenshot_dir — thư mục lưu ảnh chụp
      chromium_path  — đường dẫn Chromium (None = Playwright tự tìm)
      max_captcha    — số lần thử captcha tối đa
      page_wait      — giây chờ sau submit
      progress_cb    — callback(processed, total, yes, no, captcha_err, error)
      log_cb         — callback(str_message)
      stop_event     — threading.Event để dừng từ ngoài
    """

    def __init__(self, input_path, output_dir, screenshot_dir,
                 chromium_path=None, max_captcha=5, page_wait=2.5,
                 progress_cb=None, log_cb=None, stop_event=None):
        self.input_path     = input_path
        self.output_dir     = output_dir
        self.screenshot_dir = screenshot_dir
        self.chromium_path  = chromium_path
        self.max_captcha    = max_captcha
        self.page_wait      = page_wait
        self.progress_cb    = progress_cb or (lambda **kw: None)
        self._log_cb        = log_cb or (lambda msg: None)
        self.stop_event     = stop_event
        self.results        = []
        self.stats          = dict(yes=0, no=0, captcha_err=0, error=0)
        self.out_path       = None

        os.makedirs(output_dir, exist_ok=True)
        os.makedirs(screenshot_dir, exist_ok=True)

        if _OCR is None:
            raise RuntimeError('ddddocr không khả dụng.')

    # ── Logging ─────────────────────────────────────────────
    def _log(self, msg):
        ts = datetime.now().strftime('%H:%M:%S')
        full = f'[{ts}] {msg}'
        print(full)
        self._log_cb(full)

    # ── Output filename ──────────────────────────────────────
    def _make_output_path(self):
        today = datetime.now().strftime('%y%m%d')
        pat = re.compile(rf'^{today}_KQ_check_hoa_don_(\d+)\.xlsx$', re.I)
        max_n = 0
        try:
            for f in os.listdir(self.output_dir):
                m = pat.match(f)
                if m:
                    max_n = max(max_n, int(m.group(1)))
        except Exception:
            pass
        return os.path.join(self.output_dir,
                            f'{today}_KQ_check_hoa_don_{max_n + 1:02d}.xlsx')

    # ── Captcha ──────────────────────────────────────────────
    def _get_captcha_bytes(self, page):
        try:
            elem = page.get_by_role('img', name=re.compile(r'captcha', re.I))
            elem.wait_for(state='visible', timeout=5000)
            elem.scroll_into_view_if_needed()
            return elem.screenshot()
        except Exception:
            pass
        for sel in ["img.captcha-img", "img[src*='captcha']", '#captchaImg', '.captcha img']:
            try:
                elem = page.locator(sel).first
                elem.wait_for(state='visible', timeout=2000)
                return elem.screenshot()
            except Exception:
                pass
        return None

    def _solve_captcha(self, img_bytes):
        try:
            text = _OCR.classification(img_bytes)
            return text.strip().upper()
        except Exception as e:
            self._log(f'    [ddddocr err] {e}')
            return ''

    def _refresh_captcha(self, page):
        try:
            page.get_by_role('button').nth(4).click(timeout=2000)
            time.sleep(0.8)
            return True
        except Exception:
            pass
        for sel in ['span.refresh-captcha', '.captcha-refresh',
                    "span[ng-click*='captcha']", 'img.captcha-img']:
            try:
                page.locator(sel).first.click(timeout=1500)
                time.sleep(0.8)
                return True
            except Exception:
                pass
        return False

    # ── Popup ────────────────────────────────────────────────
    def _dismiss_popup(self, page, first_time=False):
        time.sleep(3 if first_time else 1)
        for locator in [
            page.get_by_role('button', name=re.compile(r'close|đóng|×', re.I)),
            page.locator("button.close, [data-dismiss='modal'], .btn-close").first,
        ]:
            try:
                locator.click(timeout=2500)
                self._log('    [popup] Đã đóng.')
                time.sleep(0.5)
                return
            except Exception:
                pass
        try:
            page.keyboard.press('Escape')
        except Exception:
            pass

    # ── Fill form ────────────────────────────────────────────
    def _fill_form(self, page, row):
        mst        = str(row.get('MST', '')).strip()
        ky_hieu    = str(row.get('Ky_hieu', '')).strip()
        so_hd      = str(row.get('So_HD', '')).strip()
        loai_hd    = str(row.get('Loai_HD', 'GTGT')).strip().upper()
        thanh_toan = re.sub(r'[^\d]', '', str(row.get('Thanh_toan', '')))

        page.get_by_role('textbox', name=re.compile(r'MST người bán', re.I)).fill(mst)

        try:
            page.get_by_role('combobox').first.click()
            kw = re.compile(r'giá trị gia tăng', re.I) if 'GTGT' in loai_hd \
                 else re.compile(r'bán hàng', re.I)
            page.get_by_role('option', name=kw).first.click(timeout=3000)
        except Exception as e:
            self._log(f'    [warn] Dropdown loại HĐ: {e}')

        page.get_by_role('textbox', name=re.compile(r'Ký hiệu hóa đơn', re.I)).fill(ky_hieu)
        page.get_by_role('textbox', name=re.compile(r'Số hóa đơn', re.I)).fill(so_hd)

        if thanh_toan:
            try:
                page.get_by_role('textbox', name=re.compile(r'Tổng tiền thanh toán', re.I)).fill(thanh_toan)
            except Exception:
                pass

    # ── Parse result ─────────────────────────────────────────
    def _parse_result(self, page):
        """
        QUAN TRỌNG: check 'Không tồn tại' TRƯỚC 'Tồn tại'.
        'Không tồn tại hóa đơn' chứa 'tồn tại hóa đơn' → check NO trước YES.
        """
        body = page.locator('body').inner_text()

        if re.search(r'Mã captcha không đúng|Sai mã captcha|mã xác nhận', body, re.I):
            return 'CAPTCHA_ERR', 'Captcha sai'

        if re.search(r'Không tồn tại hóa đơn', body, re.I):
            return 'NO', 'No - Không tồn tại hóa đơn'

        if re.search(r'Tồn tại hóa đơn', body, re.I):
            match = re.search(r'Trạng thái xử lý[^\n:]*[:\s]+([^\n]+)', body)
            status = match.group(1).strip() if match else 'Đã cấp mã hóa đơn'
            return 'YES', f'Yes - Tồn tại hóa đơn | {status}'

        return 'UNKNOWN', 'Không xác định kết quả'

    # ── Process one invoice ──────────────────────────────────
    def _process_one(self, page, row, stt, total):
        mst   = str(row.get('MST', '')).strip()
        so_hd = str(row.get('So_HD', '')).strip()
        self._log(f'[>] {stt}/{total}  MST={mst}  KH={row.get("Ky_hieu","")}  SHD={so_hd}')

        if self.stop_event and self.stop_event.is_set():
            return 'UNKNOWN', 'Dừng bởi người dùng', '', 0

        page.goto('https://hoadondientu.gdt.gov.vn/',
                  wait_until='networkidle', timeout=15000)
        self._dismiss_popup(page, first_time=(stt == 1))

        try:
            page.get_by_role('tab', name=re.compile(r'Tra cứu', re.I)).click(timeout=3000)
            time.sleep(0.5)
        except Exception:
            pass

        try:
            self._fill_form(page, row)
        except Exception as e:
            self._log(f'    [!] Lỗi điền form: {e}')
            return 'ERROR', f'Lỗi điền form: {e}', '', 0

        code, label, attempts = 'UNKNOWN', 'Chưa có kết quả', 0

        for attempt in range(1, self.max_captcha + 1):
            attempts = attempt
            if self.stop_event and self.stop_event.is_set():
                break

            cap_bytes = self._get_captcha_bytes(page)
            if cap_bytes:
                cap_text = self._solve_captcha(cap_bytes)
                self._log(f'    [cap #{attempt}] ddddocr → "{cap_text}"')
            else:
                cap_text = ''
                self._log(f'    [cap #{attempt}] Không lấy được ảnh captcha')

            if not cap_text:
                if attempt < self.max_captcha:
                    self._refresh_captcha(page)
                continue

            try:
                inp = page.get_by_role('textbox', name=re.compile(r'Nhập mã captcha', re.I))
                inp.scroll_into_view_if_needed()
                inp.fill(cap_text)
                page.get_by_role('button', name=re.compile(r'Tìm kiếm', re.I)).click()
                time.sleep(self.page_wait)
            except Exception as e:
                self._log(f'    [!] Submit lỗi: {e}')
                break

            code, label = self._parse_result(page)

            if code == 'CAPTCHA_ERR':
                self._log(f'    [!] Captcha sai (lần {attempt}/{self.max_captcha})')
                if attempt < self.max_captcha:
                    self._refresh_captcha(page)
                    time.sleep(0.5)
                continue

            break

        # Stats
        if code == 'YES':
            self.stats['yes'] += 1
        elif code == 'NO':
            self.stats['no'] += 1
        elif code == 'CAPTCHA_ERR':
            self.stats['captcha_err'] += 1
        else:
            self.stats['error'] += 1

        # Screenshot
        date_s = datetime.now().strftime('%y%m%d')
        mst_s  = re.sub(r'[^\w]', '', mst)
        fname  = f'{stt:03d}_{date_s}_{mst_s}_{so_hd}_{code}.png'
        fpath  = os.path.join(self.screenshot_dir, fname)
        try:
            page.screenshot(path=fpath, full_page=True)
        except Exception:
            fpath = ''

        self._log(f'    → {label}')
        return code, label, fpath if fpath and os.path.exists(fpath) else '', attempts

    # ── Export Excel ─────────────────────────────────────────
    def _export_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'KQ_Check_HoaDon'

        hdr_fill = PatternFill('solid', fgColor='1F3864')
        hdr_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
        thin     = Side(style='thin', color='CCCCCC')
        border   = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.row_dimensions[1].height = 28
        for ci, (key, header) in enumerate(OUTPUT_COLUMNS, 1):
            c = ws.cell(row=1, column=ci, value=header)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = border

        ws.freeze_panes = 'A2'

        for ri, res in enumerate(self.results, 2):
            code  = res.get('Ket_qua', 'UNKNOWN')
            rfill = PatternFill('solid', fgColor=RESULT_COLOR.get(code, 'FFFFFF'))
            ws.row_dimensions[ri].height = 18

            for ci, (key, _) in enumerate(OUTPUT_COLUMNS, 1):
                val = res.get(key, '')
                c   = ws.cell(row=ri, column=ci, value=val)
                c.fill      = rfill
                c.border    = border
                c.alignment = Alignment(vertical='center')
                c.font      = Font(name='Arial', size=10, bold=(key == 'Ket_qua'))

        for ci, (key, header) in enumerate(OUTPUT_COLUMNS, 1):
            vals  = [str(r.get(key, '')) for r in self.results]
            width = max(len(header), max((len(v) for v in vals), default=0)) + 3
            ws.column_dimensions[get_column_letter(ci)].width = min(width, 55)

        ws.auto_filter.ref = f'A1:{get_column_letter(len(OUTPUT_COLUMNS))}1'

        # Sheet tổng kết
        ws2 = wb.create_sheet('Tong_ket')
        ws2.column_dimensions['A'].width = 35
        ws2.column_dimensions['B'].width = 15
        total = len(self.results)
        rows_sum = [
            ('Tổng số hóa đơn kiểm tra', total,                                      '1F3864', 'FFFFFF'),
            ('✓ YES - Tồn tại, hợp lệ',   self.stats['yes'],                         'C6EFCE', '375623'),
            ('✗ NO  - Không tồn tại',      self.stats['no'],                          'FFC7CE', '9C0006'),
            ('⚠ Lỗi Captcha',              self.stats['captcha_err'],                 'FFEB9C', '9C6500'),
            ('! Timeout / Lỗi khác',       self.stats['error'],                       'F4B942', '000000'),
            ('Ngày giờ kiểm tra',          datetime.now().strftime('%d/%m/%Y %H:%M'), 'FFFFFF', '000000'),
            ('File kết quả',               self.out_path,                             'FFFFFF', '000000'),
        ]
        for si, (lbl, val, bg, fg) in enumerate(rows_sum, 1):
            for ci, v in enumerate([lbl, val], 1):
                c = ws2.cell(row=si, column=ci, value=v)
                c.fill      = PatternFill('solid', fgColor=bg)
                c.font      = Font(bold=True, color=fg, name='Arial')
                c.alignment = Alignment(vertical='center')

        wb.save(self.out_path)
        self._log(f'[+] Đã lưu Excel: {self.out_path}')

    # ── Zip screenshots ──────────────────────────────────────
    def zip_screenshots(self, zip_path):
        """Nén tất cả ảnh trong screenshot_dir thành 1 file zip."""
        count = 0
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for f in sorted(os.listdir(self.screenshot_dir)):
                if f.lower().endswith('.png'):
                    zf.write(os.path.join(self.screenshot_dir, f), f)
                    count += 1
        self._log(f'[+] Đã nén {count} ảnh → {zip_path}')
        return zip_path if count > 0 else None

    # ── Main run ─────────────────────────────────────────────
    def run(self):
        self.out_path = self._make_output_path()
        self._log('═' * 60)
        self._log('  KIỂM TRA HÓA ĐƠN - hoadondientu.gdt.gov.vn (WEB MODE)')
        self._log(f'  File đầu vào : {self.input_path}')
        self._log(f'  File kết quả : {self.out_path}')
        self._log('═' * 60)

        df = read_input_excel(self.input_path)
        total = len(df)
        self._log(f'[*] Đọc {total} hóa đơn (đã tự tách Ký hiệu/Số HĐ nếu file dùng cột gộp)')

        self.progress_cb(processed=0, total=total,
                         yes=0, no=0, captcha_err=0, error=0)

        launch_opts = dict(
            headless=True,
            args=['--no-sandbox', '--disable-setuid-sandbox',
                  '--disable-dev-shm-usage',
                  '--disable-blink-features=AutomationControlled']
        )
        if self.chromium_path:
            launch_opts['executable_path'] = self.chromium_path

        with sync_playwright() as p:
            browser = p.chromium.launch(**launch_opts)
            context = browser.new_context(
                viewport={'width': 1600, 'height': 900},
                user_agent=(
                    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                    'AppleWebKit/537.36 (KHTML, like Gecko) '
                    'Chrome/124.0.0.0 Safari/537.36'
                )
            )
            page = context.new_page()

            try:
                for idx, row in df.iterrows():
                    if self.stop_event and self.stop_event.is_set():
                        self._log('[!] Dừng bởi người dùng.')
                        break

                    stt = int(idx) + 1
                    code, label, fpath, n_tries = self._process_one(
                        page, row, stt, total)

                    res = {
                        'STT':            stt,
                        'Ket_qua':        code,
                        'Mo_ta':          label,
                        'So_lan_captcha': n_tries,
                        'Ngay_check':     datetime.now().strftime('%d/%m/%Y %H:%M'),
                        'Anh_chup':       fpath,
                    }
                    for col in df.columns:
                        res[col] = row.get(col, '')
                    self.results.append(res)

                    self.progress_cb(
                        processed=stt, total=total,
                        yes=self.stats['yes'], no=self.stats['no'],
                        captcha_err=self.stats['captcha_err'],
                        error=self.stats['error']
                    )

                    if stt % 3 == 0 or stt == total:
                        try:
                            self._export_excel()
                        except Exception as e:
                            self._log(f'    [warn] Lưu tạm lỗi: {e}')

            except KeyboardInterrupt:
                self._log('[!] Dừng bởi Ctrl+C.')
            finally:
                try:
                    self._export_excel()
                except Exception as e:
                    self._log(f'[!] Lỗi lưu Excel cuối: {e}')
                browser.close()

        self._log('═' * 60)
        self._log(f'  HOÀN THÀNH — {len(self.results)}/{total} hóa đơn')
        self._log(f'  ✓ YES : {self.stats["yes"]}')
        self._log(f'  ✗ NO  : {self.stats["no"]}')
        self._log(f'  ⚠ Cap : {self.stats["captcha_err"]}')
        self._log(f'  ! Err : {self.stats["error"]}')
        self._log('═' * 60)

        return self.results, self.stats, self.out_path
