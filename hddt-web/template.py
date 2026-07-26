#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
template.py — Sinh file Excel mẫu chuẩn cho người dùng tải về và điền dữ liệu.

Chỉ giữ đúng các cột cần thiết cho việc tra cứu (khác với các file cũ có quá
nhiều cột thừa). Cột "Số & Ký hiệu hóa đơn" cho phép nhập gộp kiểu
"1-C26TAP-00001765" — checker sẽ tự tách ra Ký hiệu (C26TAP) và Số HĐ
(00001765), xem input_utils.py.
"""

import io

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# (key, header, width, required)
TEMPLATE_COLUMNS = [
    ('MST',          'MST',                       16, True),
    ('Ten_Cty',      'Tên công ty',                38, False),
    ('So_KyHieu_HD', 'Số & Ký hiệu hóa đơn',       24, True),
    ('Loai_HD',      'Loại hóa đơn',               16, True),
    ('Thanh_toan',   'Tiền thanh toán',            18, True),
    ('Tien_thue',    'Tiền thuế',                  16, False),
    ('Ngay',         'Ngày hóa đơn',                14, False),
]

SAMPLE_ROWS = [
    ('0313756193', 'CÔNG TY TNHH AEON TOPVALU VIỆT NAM', '1-C26TAP-00001765', 'GTGT',      610149672, 610149672, '14/05/2026'),
    ('0900234071', 'CÔNG TY TNHH VINATECH',              '1-C26TVN-00000455', 'Ban_hang',  745447700, None,      '04/05/2026'),
]

HEADER_FILL = PatternFill('solid', fgColor='1F3864')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
REQUIRED_FILL = PatternFill('solid', fgColor='FFF2CC')
THIN = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATA_ROWS_RESERVED = 500  # số dòng áp dropdown/định dạng sẵn cho người dùng điền


def _build_data_sheet(wb):
    ws = wb.active
    ws.title = 'Danh_sach_hoa_don'

    # QUAN TRỌNG: tiêu đề dòng 1 PHẢI đúng bằng tên cột kỹ thuật (key, vd
    # "So_KyHieu_HD") vì code đọc file bằng đúng tên này (xem input_utils.py,
    # checker_web.py) — không phải nhãn tiếng Việt đẹp. Nhãn tiếng Việt +
    # dấu (*) bắt buộc đặt trong comment của ô, và trong sheet "Huong_dan".
    ws.row_dimensions[1].height = 30
    for ci, (key, header, width, required) in enumerate(TEMPLATE_COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=key)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
        c.comment = Comment(
            f'{header}{" (bắt buộc)" if required else " (tùy chọn)"}',
            'HDDT Checker', width=200, height=40)
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.cell(row=1, column=3).comment = Comment(
        'Số & Ký hiệu hóa đơn (bắt buộc)\n'
        'Nhập gộp theo mẫu: <mẫu số>-<ký hiệu>-<số hóa đơn>\n'
        'Ví dụ: 1-C26TAP-00001765\n'
        '→ Hệ thống tự tách: Ký hiệu = C26TAP, Số HĐ = 00001765.\n'
        'Không cần tách sẵn ra 2 cột.', 'HDDT Checker', width=260, height=120)
    ws.cell(row=1, column=4).comment = Comment(
        'Loại hóa đơn (bắt buộc)\n'
        'Chỉ nhận 1 trong 2 giá trị:\n'
        '  GTGT       = Hóa đơn điện tử giá trị gia tăng\n'
        '  Ban_hang   = Hóa đơn điện tử bán hàng', 'HDDT Checker', width=260, height=90)
    ws.cell(row=1, column=5).comment = Comment(
        'Tiền thanh toán (bắt buộc)\n'
        'Số nguyên, không dấu chấm/phẩy. Ví dụ: 746618472', 'HDDT Checker', width=220, height=60)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(TEMPLATE_COLUMNS))}1'

    # Dữ liệu mẫu (2 dòng minh họa, có thể xóa trước khi điền thật)
    for ri, sample in enumerate(SAMPLE_ROWS, 2):
        for ci, val in enumerate(sample, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = BORDER
            c.font = Font(name='Arial', size=10, italic=True, color='808080')
            c.alignment = Alignment(vertical='center')

    # Tô nhạt các cột bắt buộc cho toàn bộ vùng nhập liệu dự phòng
    last_row = 1 + DATA_ROWS_RESERVED
    for ci, (key, header, width, required) in enumerate(TEMPLATE_COLUMNS, 1):
        if not required:
            continue
        for ri in range(len(SAMPLE_ROWS) + 2, last_row + 1):
            ws.cell(row=ri, column=ci).fill = REQUIRED_FILL
            ws.cell(row=ri, column=ci).border = BORDER

    # Dropdown cho cột Loại hóa đơn
    loai_col_letter = get_column_letter([k for k, *_ in TEMPLATE_COLUMNS].index('Loai_HD') + 1)
    dv = DataValidation(type='list', formula1='"GTGT,Ban_hang"', allow_blank=True,
                         showDropDown=False, showErrorMessage=True,
                         errorTitle='Giá trị không hợp lệ',
                         error='Chỉ được chọn GTGT hoặc Ban_hang.')
    ws.add_data_validation(dv)
    dv.add(f'{loai_col_letter}2:{loai_col_letter}{last_row}')

    return ws


def _build_instruction_sheet(wb):
    ws = wb.create_sheet('Huong_dan')
    ws.column_dimensions['A'].width = 100
    ws.sheet_view.showGridLines = False

    lines = [
        ('HƯỚNG DẪN ĐIỀN FILE EXCEL ĐẦU VÀO', True, '1F3864', 'FFFFFF', 14),
        ('', False, None, None, 11),
        ('1. Xóa 2 dòng dữ liệu mẫu (in nghiêng, màu xám) trước khi điền dữ liệu thật.', False, None, None, 11),
        ('2. Mỗi dòng = 1 hóa đơn cần tra cứu trên hoadondientu.gdt.gov.vn.', False, None, None, 11),
        ('3. Các cột có dấu (*) ở tiêu đề là BẮT BUỘC phải điền: MST, Số & Ký hiệu hóa đơn, '
         'Loại hóa đơn, Tiền thanh toán.', False, None, None, 11),
        ('', False, None, None, 11),
        ('Cột "Số & Ký hiệu hóa đơn":', True, None, None, 12),
        ('  Nhập gộp đúng theo mẫu trên hóa đơn: <mẫu số>-<ký hiệu>-<số hóa đơn>', False, None, None, 11),
        ('  Ví dụ: 1-C26TAP-00001765', False, None, None, 11),
        ('  → Hệ thống tự tách: Ký hiệu hóa đơn = C26TAP, Số hóa đơn = 00001765.', False, None, None, 11),
        ('  Không cần tự tách sẵn ra 2 cột riêng.', False, None, None, 11),
        ('', False, None, None, 11),
        ('Cột "Loại hóa đơn": chỉ nhận 1 trong 2 giá trị (đã có sẵn dropdown để chọn):', True, None, None, 12),
        ('  GTGT       = Hóa đơn điện tử giá trị gia tăng', False, None, None, 11),
        ('  Ban_hang   = Hóa đơn điện tử bán hàng', False, None, None, 11),
        ('', False, None, None, 11),
        ('Cột "Tiền thanh toán" / "Tiền thuế": số nguyên, KHÔNG có dấu chấm/phẩy phân cách '
         '(ví dụ 746618472, không phải 746.618.472).', False, None, None, 11),
        ('', False, None, None, 11),
        ('Cột "Tên công ty" và "Ngày hóa đơn": không bắt buộc, chỉ để tham khảo/đối chiếu, '
         'không ảnh hưởng đến kết quả tra cứu.', False, None, None, 11),
    ]
    for ri, (text, bold, bg, fg, size) in enumerate(lines, 1):
        c = ws.cell(row=ri, column=1, value=text)
        c.font = Font(bold=bold, name='Arial', size=size, color=fg or '000000')
        c.alignment = Alignment(vertical='center', wrap_text=True)
        if bg:
            c.fill = PatternFill('solid', fgColor=bg)
            ws.row_dimensions[ri].height = 26


def build_template_bytes():
    """Trả về (bytes) nội dung file Excel mẫu, sẵn sàng để gửi qua Flask send_file."""
    wb = openpyxl.Workbook()
    _build_data_sheet(wb)
    _build_instruction_sheet(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
