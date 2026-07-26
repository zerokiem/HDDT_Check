#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════╗
║   KIỂM TRA HÓA ĐƠN ĐIỆN TỬ - hoadondientu.gdt.gov.vn       ║
║   Phiên bản: v8                                              ║
║   Captcha : ddddocr (thay thế model Keras cũ không chính xác)║
╚══════════════════════════════════════════════════════════════╝

CÀI ĐẶT:
    pip install playwright ddddocr openpyxl pandas pillow
    playwright install chromium

CẤU TRÚC THƯ MỤC:
    check_hddt_cl_v8.py    ← script này
    Hoa_don_input.xlsx     ← file dữ liệu đầu vào
    screenshots/           ← tự tạo
"""

import os
import re
import sys
import time
import pandas as pd
from datetime import datetime

# ── Suppress TF/oneDNN noise ────────────────────────────────
os.environ['TF_CPP_MIN_LOG_LEVEL']      = '3'
os.environ['TF_ENABLE_ONEDNN_OPTS']     = '0'
os.environ['TF_USE_LEGACY_KERAS']       = '1'
import warnings
warnings.filterwarnings('ignore')

import logging
logging.getLogger('tensorflow').setLevel(logging.FATAL)
logging.getLogger('absl').setLevel(logging.FATAL)

# ─── CẤU HÌNH ────────────────────────────────────────────────
EXCEL_INPUT      = "Hoa_don_input.xlsx"
OUTPUT_DIR       = "."           # Thư mục lưu file kết quả
SCREENSHOT_DIR   = "screenshots" # Thư mục lưu ảnh
MAX_CAPTCHA_AUTO = 5             # Số lần ddddocr tự đoán
MANUAL_FALLBACK  = False         # Hỏi tay nếu vẫn sai
PAGE_WAIT        = 2.5           # Giây chờ sau khi click Tìm kiếm
RELOAD_BETWEEN   = True          # Reload trang mỗi hóa đơn
# ─────────────────────────────────────────────────────────────

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PwTimeout
except ImportError:
    print("[!] Thiếu playwright. Chạy: pip install playwright && playwright install chromium")
    sys.exit(1)

try:
    import ddddocr
    _ocr = ddddocr.DdddOcr(show_ad=False)
    print("[+] ddddocr sẵn sàng.")
except ImportError:
    print("[!] Thiếu ddddocr. Chạy: pip install ddddocr")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[!] Thiếu openpyxl. Chạy: pip install openpyxl")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════
# HELPER: TÊN FILE OUTPUT TỰ TĂNG
# ══════════════════════════════════════════════════════════════
def make_output_path(base_dir=OUTPUT_DIR):
    """
    Tên file: yymmdd_KQ_check_hoa_don_NNN.xlsx
    Nếu file đã tồn tại → tăng NNN lên.
    """
    today = datetime.now().strftime('%y%m%d')
    pattern = re.compile(rf'^{today}_KQ_check_hoa_don_(\d+)\.xlsx$', re.I)
    max_n = 0
    try:
        for f in os.listdir(base_dir):
            m = pattern.match(f)
            if m:
                max_n = max(max_n, int(m.group(1)))
    except Exception:
        pass
    return os.path.join(base_dir, f"{today}_KQ_check_hoa_don_{max_n + 1:02d}.xlsx")


# ══════════════════════════════════════════════════════════════
# CAPTCHA SOLVER
# ══════════════════════════════════════════════════════════════
def solve_captcha_auto(img_bytes: bytes) -> str:
    """
    Dùng ddddocr đọc captcha từ PNG bytes (Playwright screenshot).
    Trả về chuỗi uppercase, hoặc "" nếu thất bại.
    """
    try:
        text = _ocr.classification(img_bytes)
        return text.strip().upper()
    except Exception as e:
        print(f"    [ddddocr err] {e}")
        return ""


def get_captcha_bytes(page) -> bytes | None:
    """Chụp ảnh element captcha từ trang."""
    try:
        elem = page.get_by_role("img", name=re.compile(r"captcha", re.I))
        elem.wait_for(state="visible", timeout=5000)
        elem.scroll_into_view_if_needed()
        return elem.screenshot()
    except Exception:
        pass
    # Fallback selectors nếu role không match
    for sel in ["img.captcha-img", "img[src*='captcha']", "#captchaImg", ".captcha img"]:
        try:
            elem = page.locator(sel).first
            elem.wait_for(state="visible", timeout=2000)
            return elem.screenshot()
        except Exception:
            pass
    return None


def save_captcha_png(img_bytes: bytes, path: str):
    with open(path, 'wb') as f:
        f.write(img_bytes)
    try:
        import subprocess
        if sys.platform == "win32":
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception:
        pass


def refresh_captcha(page):
    """Click vào ảnh captcha để đổi mã mới."""
    try:
        #page.get_by_role("img", name=re.compile(r"captcha", re.I)).click(timeout=2000)
        page.get_by_role("button").nth(4).click(timeout=2000)	# đổi captcha
        time.sleep(0.8)
        return True
    except Exception:
        pass
    for sel in ["span.refresh-captcha", ".captcha-refresh",
                "span[ng-click*='captcha']", "img.captcha-img"]:
        try:
            page.locator(sel).first.click(timeout=1500)
            time.sleep(0.8)
            return True
        except Exception:
            pass
    return False


# ══════════════════════════════════════════════════════════════
# POPUP
# ══════════════════════════════════════════════════════════════
def dismiss_popup(page, first_time=False):
    wait_s = 3 if first_time else 1
    time.sleep(wait_s)
    for locator in [
        page.get_by_role("button", name=re.compile(r"close|đóng|×", re.I)),
        page.locator("button.close, [data-dismiss='modal'], .btn-close").first,
    ]:
        try:
            locator.click(timeout=2500)
            print("    [popup] Đã đóng.")
            time.sleep(0.5)
            return
        except Exception:
            pass
    try:
        page.keyboard.press("Escape")
    except Exception:
        pass


# ══════════════════════════════════════════════════════════════
# ĐIỀN FORM
# ══════════════════════════════════════════════════════════════
def fill_form(page, row):
    mst        = str(row.get('MST', '')).strip()
    ky_hieu    = str(row.get('Ky_hieu', '')).strip()
    so_hd      = str(row.get('So_HD', '')).strip()
    loai_hd    = str(row.get('Loai_HD', 'GTGT')).strip().upper()
    thanh_toan = re.sub(r'[^\d]', '', str(row.get('Thanh_toan', '')))

    # 1. MST
    page.get_by_role("textbox", name=re.compile(r"MST người bán", re.I)).fill(mst)

    # 2. Loại HĐ dropdown
    try:
        page.get_by_role("combobox").first.click()
        kw = re.compile(r"giá trị gia tăng", re.I) if "GTGT" in loai_hd \
             else re.compile(r"bán hàng", re.I)
        page.get_by_role("option", name=kw).first.click(timeout=3000)
    except Exception as e:
        print(f"    [warn] Dropdown loại HĐ: {e}")

    # 3. Ký hiệu
    page.get_by_role("textbox", name=re.compile(r"Ký hiệu hóa đơn", re.I)).fill(ky_hieu)

    # 4. Số HĐ
    page.get_by_role("textbox", name=re.compile(r"Số hóa đơn", re.I)).fill(so_hd)

    # 5. Tổng tiền thanh toán
    if thanh_toan:
        try:
            page.get_by_role("textbox", name=re.compile(r"Tổng tiền thanh toán", re.I)).fill(thanh_toan)
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════
# ĐỌC KẾT QUẢ
# ══════════════════════════════════════════════════════════════
def parse_result(page):
    """
    Trả về: (code, label)
      'CAPTCHA_ERR' | 'YES' | 'NO' | 'UNKNOWN'

    QUAN TRỌNG: Phải check 'Không tồn tại' TRƯỚC 'Tồn tại'.
    Lý do: chuỗi "Không tồn tại hóa đơn có thông tin trùng khớp"
    chứa luôn "tồn tại hóa đơn có thông tin trùng khớp" → nếu check YES
    trước thì regex sẽ khớp sai cả hai trường hợp.
    """
    body = page.locator("body").inner_text()

    # 1. Captcha sai
    if re.search(r"Mã captcha không đúng|Sai mã captcha|mã xác nhận", body, re.I):
        return 'CAPTCHA_ERR', "Captcha sai"

    # 2. KHÔNG tồn tại → phải check TRƯỚC khi check tồn tại
    if re.search(r"Không tồn tại hóa đơn", body, re.I):
        return 'NO', "No - Không tồn tại hóa đơn"

    # 3. Tồn tại (chỉ đến đây nếu KHÔNG có chữ "Không tồn tại")
    if re.search(r"Tồn tại hóa đơn", body, re.I):
        match = re.search(r"Trạng thái xử lý[^\n:]*[:\s]+([^\n]+)", body)
        status = match.group(1).strip() if match else "Đã cấp mã hóa đơn"
        return 'YES', f"Yes - Tồn tại hóa đơn | {status}"

    return 'UNKNOWN', "Không xác định kết quả"


# ══════════════════════════════════════════════════════════════
# XUẤT EXCEL ĐẸP
# ══════════════════════════════════════════════════════════════
RESULT_COLOR = {
    'YES'         : 'C6EFCE',
    'NO'          : 'FFC7CE',
    'CAPTCHA_ERR' : 'FFEB9C',
    'UNKNOWN'     : 'D9D9D9',
    'ERROR'       : 'F4B942',
}

# Thứ tự cột xuất ra Excel: (key trong dict result, tên hiển thị header)
OUTPUT_COLUMNS = [
    ("STT",             "STT"),
    ("MST",             "MST"),
    ("Ten_Cty",         "Tên công ty"),
    ("So_HD",           "Số hóa đơn"),
    ("Ky_hieu",         "Ký hiệu HĐ"),
    ("Tien_thue",       "Tiền thuế"),
    ("Thanh_toan",      "Tiền thanh toán"),
    ("Loai_HD",         "Loại hóa đơn"),
    ("Ngay",            "Ngày hóa đơn"),
    ("Ngay_check",      "Ngày kiểm tra"),
    ("Ket_qua",         "Kết quả"),
    ("Mo_ta",           "Mô tả kết quả"),
    ("Anh_chup",        "Ảnh chụp"),
    ("So_lan_captcha",  "Số lần captcha"),
]

def export_excel(df_orig, results: list, out_path: str):
    """Xuất file Excel với màu sắc theo kết quả, đúng thứ tự cột."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KQ_Check_HoaDon"

    hdr_fill  = PatternFill("solid", fgColor="1F3864")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Header ────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    for ci, (key, header) in enumerate(OUTPUT_COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=header)
        c.fill      = hdr_fill
        c.font      = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border

    ws.freeze_panes = "A2"

    # ── Data rows ─────────────────────────────────────────────
    for ri, res in enumerate(results, 2):
        code  = res.get("Ket_qua", "UNKNOWN")
        rfill = PatternFill("solid", fgColor=RESULT_COLOR.get(code, "FFFFFF"))
        ws.row_dimensions[ri].height = 18

        for ci, (key, _) in enumerate(OUTPUT_COLUMNS, 1):
            val = res.get(key, "")
            c   = ws.cell(row=ri, column=ci, value=val)
            c.fill      = rfill
            c.border    = border
            c.alignment = Alignment(vertical="center")
            c.font      = Font(name="Arial", size=10, bold=(key == "Ket_qua"))

        # Hyperlink cho cột Ảnh chụp
        anh_ci = next(ci for ci, (k, _) in enumerate(OUTPUT_COLUMNS, 1) if k == "Anh_chup")
        path   = res.get("Anh_chup", "")
        if path and os.path.exists(path):
            lnk = path.replace("\\", "/")
            ws.cell(row=ri, column=anh_ci).hyperlink = f"file:///{lnk}"
            ws.cell(row=ri, column=anh_ci).font = Font(
                color="0000FF", underline="single", name="Arial", size=10)

    # ── Auto-width ────────────────────────────────────────────
    for ci, (key, header) in enumerate(OUTPUT_COLUMNS, 1):
        vals  = [str(r.get(key, "")) for r in results]
        width = max(len(header), max((len(v) for v in vals), default=0)) + 3
        ws.column_dimensions[get_column_letter(ci)].width = min(width, 55)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_COLUMNS))}1"

    # ── Sheet tổng kết ────────────────────────────────────────
    ws2 = wb.create_sheet("Tong_ket")
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 15
    total  = len(results)
    yes_n  = sum(1 for r in results if r.get("Ket_qua") == "YES")
    no_n   = sum(1 for r in results if r.get("Ket_qua") == "NO")
    cap_n  = sum(1 for r in results if r.get("Ket_qua") == "CAPTCHA_ERR")
    err_n  = total - yes_n - no_n - cap_n
    rows_sum = [
        ("Tổng số hóa đơn kiểm tra", total,                                    "1F3864", "FFFFFF"),
        ("✓ YES - Tồn tại, hợp lệ",  yes_n,                                   "C6EFCE", "375623"),
        ("✗ NO  - Không tồn tại",    no_n,                                     "FFC7CE", "9C0006"),
        ("⚠ Lỗi Captcha",            cap_n,                                    "FFEB9C", "9C6500"),
        ("! Timeout / Lỗi khác",     err_n,                                    "F4B942", "000000"),
        ("Ngày giờ kiểm tra",        datetime.now().strftime("%d/%m/%Y %H:%M"), "FFFFFF", "000000"),
        ("File kết quả",             out_path,                                  "FFFFFF", "000000"),
    ]
    for si, (lbl, val, bg, fg) in enumerate(rows_sum, 1):
        for ci, v in enumerate([lbl, val], 1):
            c = ws2.cell(row=si, column=ci, value=v)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.font      = Font(bold=True, color=fg, name="Arial")
            c.alignment = Alignment(vertical="center")

    wb.save(out_path)
    print(f"[+] Đã lưu: {os.path.abspath(out_path)}")
    return out_path


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════
class InvoiceChecker:
    def __init__(self):
        os.makedirs(SCREENSHOT_DIR, exist_ok=True)
        self.out_path = make_output_path(OUTPUT_DIR)
        self.stats = dict(yes=0, no=0, captcha_fail=0, error=0)

    def process_one(self, page, row, stt, total):
        mst   = str(row.get('MST', '')).strip()
        so_hd = str(row.get('So_HD', '')).strip()
        print(f"\n{'─'*60}")
        print(f"[>] {stt}/{total}  MST={mst}  KH={row.get('Ky_hieu','')}  SHD={so_hd}")

        if RELOAD_BETWEEN:
            page.goto("https://hoadondientu.gdt.gov.vn/",
                      wait_until="networkidle", timeout=15000)
            dismiss_popup(page, first_time=(stt == 1))

        # Chuyển sang tab Tra cứu
        try:
            page.get_by_role("tab", name=re.compile(r"Tra cứu", re.I)).click(timeout=3000)
            time.sleep(0.5)
        except Exception:
            pass

        # Điền form
        try:
            fill_form(page, row)
        except Exception as e:
            print(f"    [!] Lỗi điền form: {e}")
            return "ERROR", "Lỗi điền form", "", 0

        # ── Vòng lặp captcha ──────────────────────────────────
        code = "UNKNOWN"
        label = "Chưa có kết quả"
        attempts = 0

        for attempt in range(1, MAX_CAPTCHA_AUTO + 1):
            attempts = attempt

            # Lấy ảnh captcha và đoán
            cap_bytes = get_captcha_bytes(page)
            if cap_bytes:
                cap_text = solve_captcha_auto(cap_bytes)
                print(f"    [cap #{attempt}] ddddocr → \"{cap_text}\"")
            else:
                cap_text = ""
                print(f"    [cap #{attempt}] Không lấy được ảnh captcha")

            # Nếu model trả rỗng → fallback manual ngay
            if not cap_text:
                if MANUAL_FALLBACK:
                    tmp = f"captcha_{stt:03d}_a{attempt}.png"
                    if cap_bytes:
                        save_captcha_png(cap_bytes, tmp)
                    cap_text = input(f"    → Nhập captcha thủ công (HD#{stt} lần{attempt}): ").strip().upper()
                if not cap_text:
                    continue

            # Nhập captcha vào form
            try:
                inp = page.get_by_role("textbox", name=re.compile(r"Nhập mã captcha", re.I))
                inp.scroll_into_view_if_needed()
                inp.fill(cap_text)
                page.get_by_role("button", name=re.compile(r"Tìm kiếm", re.I)).click()
                time.sleep(PAGE_WAIT)
            except Exception as e:
                print(f"    [!] Submit lỗi: {e}")
                break

            # Đọc kết quả
            code, label = parse_result(page)

            if code == 'CAPTCHA_ERR':
                print(f"    [!] Captcha sai (lần {attempt}/{MAX_CAPTCHA_AUTO})")
                if attempt < MAX_CAPTCHA_AUTO:
                    refresh_captcha(page)
                    time.sleep(0.5)
                    continue
                # Hết lần tự đoán → hỏi tay
                if MANUAL_FALLBACK:
                    print(f"\n    *** MODEL SAI {MAX_CAPTCHA_AUTO} LẦN — Nhập thủ công ***")
                    for manual_try in range(1, 4):
                        cap_bytes2 = get_captcha_bytes(page)
                        if cap_bytes2:
                            tmp2 = f"captcha_{stt:03d}_manual{manual_try}.png"
                            save_captcha_png(cap_bytes2, tmp2)
                            print(f"    Ảnh: {os.path.abspath(tmp2)}")
                        cap_text2 = input(f"    → Nhập captcha (lần {manual_try}/3): ").strip().upper()
                        if not cap_text2:
                            break
                        try:
                            inp2 = page.get_by_role("textbox", name=re.compile(r"Nhập mã captcha", re.I))
                            inp2.fill(cap_text2)
                            page.get_by_role("button", name=re.compile(r"Tìm kiếm", re.I)).click()
                            time.sleep(PAGE_WAIT)
                        except Exception:
                            break
                        code2, label2 = parse_result(page)
                        if code2 != 'CAPTCHA_ERR':
                            code, label = code2, label2
                            attempts += manual_try
                            break
                        refresh_captcha(page)
                        time.sleep(0.5)
                    else:
                        code, label = 'CAPTCHA_ERR', f"LỖI CAPTCHA (Vượt {MAX_CAPTCHA_AUTO} lần auto + 3 lần tay)"
                break

            # Có kết quả rõ ràng
            break

        # ── Cập nhật stats ────────────────────────────────────
        if code == 'YES':             self.stats['yes'] += 1
        elif code == 'NO':           self.stats['no'] += 1
        elif code == 'CAPTCHA_ERR':  self.stats['captcha_fail'] += 1
        else:                        self.stats['error'] += 1

        # ── Chụp màn hình ─────────────────────────────────────
        date_s = datetime.now().strftime("%y%m%d")
        mst_s  = re.sub(r'[^\w]', '', mst)
        fname  = f"{date_s}_{stt:03d}_{mst_s}_{so_hd}_{code}.png"
        fpath  = os.path.join(SCREENSHOT_DIR, fname)
        try:
            page.screenshot(path=fpath, full_page=True)
        except Exception:
            fpath = ""

        print(f"    → {label}")
        return code, label, os.path.abspath(fpath) if fpath and os.path.exists(fpath) else "", attempts

    def run(self):
        print(f"\n{'═'*60}")
        print(f"  KIỂM TRA HÓA ĐƠN - hoadondientu.gdt.gov.vn")
        print(f"  File đầu vào : {EXCEL_INPUT}")
        print(f"  File kết quả : {self.out_path}")
        print(f"  Captcha      : ddddocr auto (fallback: nhập tay)")
        print("═" * 60)

        if not os.path.exists(EXCEL_INPUT):
            print(f"[!] Không tìm thấy: {EXCEL_INPUT}")
            sys.exit(1)

        df = pd.read_excel(EXCEL_INPUT, dtype=str).fillna("")
        total = len(df)
        print(f"[*] Đọc {total} hóa đơn từ {EXCEL_INPUT}")

        results = []

        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=False,
                args=["--start-maximized",
                      "--disable-blink-features=AutomationControlled"]
            )
            context = browser.new_context(
                viewport={"width": 1600, "height": 900},
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                )
            )
            page = context.new_page()

            # Load trang lần đầu
            page.goto("https://hoadondientu.gdt.gov.vn/",
                      wait_until="networkidle", timeout=15000)
            dismiss_popup(page, first_time=True)

            try:
                for idx, row in df.iterrows():
                    stt = int(idx) + 1
                    code, label, fpath, n_tries = self.process_one(
                        page, row, stt, total
                    )
                    res = {
                        "STT"            : stt,
                        "Ket_qua"        : code,
                        "Mo_ta"          : label,
                        "So_lan_captcha" : n_tries,
                        "Ngay_check"     : datetime.now().strftime("%d/%m/%Y %H:%M"),
                        "Anh_chup"       : fpath,
                        "_orig"          : row.to_dict(),
                    }
                    # Copy dữ liệu gốc phẳng ra
                    for col in df.columns:
                        res[col] = row.get(col, "")
                    results.append(res)

                    # Lưu tạm mỗi 3 hóa đơn
                    if stt % 3 == 0 or stt == total:
                        try:
                            export_excel(df, results, self.out_path)
                        except Exception as e:
                            print(f"    [warn] Lưu tạm lỗi: {e}")

            except KeyboardInterrupt:
                print("\n[!] Dừng bởi người dùng (Ctrl+C).")
            finally:
                try:
                    export_excel(df, results, self.out_path)
                except Exception as e:
                    print(f"[!] Lỗi lưu Excel cuối: {e}")
                browser.close()

        # ── Tổng kết ──────────────────────────────────────────
        print(f"\n{'═'*60}")
        print(f"  HOÀN THÀNH — {len(results)}/{total} hóa đơn")
        print(f"{'─'*60}")
        print(f"  ✓ YES (hợp lệ)       : {self.stats['yes']}")
        print(f"  ✗ NO (không tồn tại): {self.stats['no']}")
        print(f"  ⚠ Lỗi Captcha       : {self.stats['captcha_fail']}")
        print(f"  ! Lỗi khác          : {self.stats['error']}")
        print(f"{'─'*60}")
        print(f"  File kết quả: {os.path.abspath(self.out_path)}")
        print(f"  Ảnh chụp    : {os.path.abspath(SCREENSHOT_DIR)}/")
        print("═" * 60)


if __name__ == "__main__":
    InvoiceChecker().run()
